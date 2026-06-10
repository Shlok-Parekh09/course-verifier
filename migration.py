"""
Migration script to convert CSV and Excel files to SQLite database
Run this once to initialize the database from existing data files
"""

import os
import sys
from db_manager import DatabaseManager

def migrate_qs_rankings():
    """Load QS rankings from CSV to SQLite"""
    csv_file = "qs_ranked.csv"
    
    if not os.path.exists(csv_file):
        print(f"[!] {csv_file} not found, skipping...")
        return
    
    print(f"[→] Migrating QS rankings from {csv_file}...")
    count = 0
    
    try:
        with open(csv_file, 'r', encoding='utf-8', errors='ignore') as f:
            for i, line in enumerate(f):
                if i == 0:  # Skip header
                    continue
                
                line = line.strip()
                if not line:
                    continue
                
                # Parse line: "Rank,University Name" or just "University Name"
                parts = line.split(',', 1)
                university_name = parts[-1].strip()
                rank_position = None
                
                if len(parts) > 1 and parts[0].isdigit():
                    rank_position = int(parts[0])
                
                DatabaseManager.insert_qs_ranking(university_name, rank_position)
                count += 1
                
                if count % 100 == 0:
                    print(f"  ✓ Inserted {count} QS rankings...")
        
        print(f"[✓] Successfully migrated {count} QS rankings to SQLite")
    
    except Exception as e:
        print(f"[!] Error migrating QS rankings: {e}")


def migrate_nirf_rankings():
    """Load NIRF rankings from CSV to SQLite"""
    csv_file = "nirf_ranked.csv"
    
    if not os.path.exists(csv_file):
        print(f"[!] {csv_file} not found, skipping...")
        return
    
    print(f"[→] Migrating NIRF rankings from {csv_file}...")
    count = 0
    
    try:
        with open(csv_file, 'r', encoding='utf-8', errors='ignore') as f:
            for i, line in enumerate(f):
                if i == 0:  # Skip header
                    continue
                
                line = line.strip()
                if not line:
                    continue
                
                # Parse line: "Rank,University Name,Category" or variants
                parts = [p.strip() for p in line.split(',')]
                
                university_name = parts[0] if len(parts) > 0 else ""
                rank_position = None
                rank_category = None
                
                if len(parts) > 1:
                    # Try to parse second column as rank
                    if parts[1].isdigit():
                        rank_position = int(parts[1])
                    else:
                        university_name = f"{parts[0]}, {parts[1]}"
                
                if len(parts) > 2:
                    rank_category = parts[2]
                
                if university_name:
                    DatabaseManager.insert_nirf_ranking(university_name, rank_position, rank_category)
                    count += 1
                
                if count % 100 == 0:
                    print(f"  ✓ Inserted {count} NIRF rankings...")
        
        print(f"[✓] Successfully migrated {count} NIRF rankings to SQLite")
    
    except Exception as e:
        print(f"[!] Error migrating NIRF rankings: {e}")


def migrate_combined_work_excel():
    """Load CombinedWork.xlsx into SQLite"""
    excel_file = "CombinedWork.xlsx"
    
    if not os.path.exists(excel_file):
        print(f"[!] {excel_file} not found, skipping...")
        return
    
    print(f"[→] Migrating course data from {excel_file}...")
    
    try:
        import openpyxl
        from openpyxl import load_workbook
    except ImportError:
        print("[!] openpyxl not installed. Install with: pip install openpyxl")
        return
    
    try:
        workbook = load_workbook(excel_file)
        sheet = workbook.active
        
        count = 0
        for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
            if row_idx == 1:  # Skip header
                continue
            
            if not any(row):  # Skip empty rows
                continue
            
            try:
                # Adjust column indices based on your Excel structure
                course_title = row[0] if len(row) > 0 else None
                university_name = row[1] if len(row) > 1 else None
                fee = row[2] if len(row) > 2 else None
                duration = row[3] if len(row) > 3 else None
                mode = row[4] if len(row) > 4 else None
                skills = row[5] if len(row) > 5 else None
                url = row[6] if len(row) > 6 else None
                
                # Convert fee to float if possible
                if fee and isinstance(fee, str):
                    fee = float(''.join(c for c in fee if c.isdigit() or c == '.'))
                
                # Convert duration to int if possible
                if duration and isinstance(duration, str):
                    duration = int(''.join(c for c in duration if c.isdigit()))
                
                conn = DatabaseManager.get_connection()
                cursor = conn.cursor()
                cursor.execute('''
                    INSERT INTO course_data (course_title, university_name, fee, duration_months, mode, skills, url)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                ''', (course_title, university_name, fee, duration, mode, skills, url))
                conn.commit()
                
                count += 1
                if count % 50 == 0:
                    print(f"  ✓ Inserted {count} course records...")
            
            except Exception as e:
                print(f"  [!] Error on row {row_idx}: {e}")
        
        print(f"[✓] Successfully migrated {count} course records to SQLite")
    
    except Exception as e:
        print(f"[!] Error migrating CombinedWork.xlsx: {e}")


def main():
    """Run all migrations"""
    print("\n" + "="*60)
    print("   Course Verifier Database Migration")
    print("="*60 + "\n")
    
    # Initialize database schema
    DatabaseManager.initialize_db()
    
    # Run migrations
    print("\n[PHASE 1] Migrating QS Rankings...")
    migrate_qs_rankings()
    
    print("\n[PHASE 2] Migrating NIRF Rankings...")
    migrate_nirf_rankings()
    
    print("\n[PHASE 3] Migrating Course Data...")
    migrate_combined_work_excel()
    
    # Print summary
    print("\n" + "="*60)
    print("   Migration Summary")
    print("="*60)
    qs_count = DatabaseManager.get_table_count('qs_rankings')
    nirf_count = DatabaseManager.get_table_count('nirf_rankings')
    course_count = DatabaseManager.get_table_count('course_data')
    
    print(f"QS Rankings:      {qs_count:,} records")
    print(f"NIRF Rankings:    {nirf_count:,} records")
    print(f"Course Data:      {course_count:,} records")
    print("="*60 + "\n")
    
    print("[✓] Migration complete! Database is ready for use.")
    print(f"[✓] Database file: {DatabaseManager.DB_PATH}")


if __name__ == "__main__":
    main()
